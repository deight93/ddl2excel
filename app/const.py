from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_ROW = 4
BASE_COL = 2

FILL_YELLOW = PatternFill("solid", fgColor="FFF000")
FILL_GRAY = PatternFill("solid", fgColor="E6E6E6")
FILL_HEADER = PatternFill("solid", fgColor="C0C0C0")
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")

BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD_FONT = Font(bold=True)
BOLD_LARGE_FONT = Font(bold=True, size=13)


META_FIELDS_KO = [
    "시스템/서비스",
    "작성자",
    "작성일",
    "프로젝트명",
    "DBMS",
    "DB 이름",
    "스키마명",
    "테이블명",
    "상세설명",
]
META_FIELDS_EN = [
    "System/Service",
    "Author",
    "Date",
    "Project",
    "DBMS",
    "DB Name",
    "Schema Name",
    "Table Name",
    "Description",
]

COLUMN_HEADERS_KO = [
    "번호",
    "컬럼명",
    "타입",
    "길이",
    "기본키",
    "Null 불가",
    "기본값",
    "정의/설명",
    "",
    "",
    "참조테이블",
    "비고",
]
COLUMN_HEADERS_EN = [
    "no",
    "Column Name",
    "Type",
    "Length",
    "PK",
    "Not Null",
    "Default",
    "Description",
    "",
    "",
    "Reference Table",
    "Note",
]

COLUMN_WIDTHS = [5, 25, 18, 8, 8, 10, 12, 28, 2, 2, 14, 10]

# 병합 위치/인덱스 헤더
COLUMN_MERGE_SPECS = {
    "desc": (8, 10),
    "index_name": (2, 4),
    "index_type": (5, 7),
    "index_unique": (8, 10),
    "index_columns": (11, 12),
}
INDEX_HEADERS_KO = [
    "번호",
    "인덱스 이름",
    "",
    "",
    "인덱스 타입",
    "",
    "",
    "유니크",
    "",
    "",
    "구성 컬럼",
    "",
]
INDEX_HEADERS_EN = [
    "no",
    "Index Name",
    "",
    "",
    "Index Type",
    "",
    "",
    "Unique",
    "",
    "",
    "Columns",
    "",
]

DATA_PERIOD_HEADERS_KO = [
    "초기 예상 건수",
    "증가 예상 건수",
    "최대 예상 건수",
    "data 보존 기간",
    "data 백업 주기",
]
DATA_PERIOD_HEADERS_EN = [
    "Initial Count",
    "Expected Growth",
    "Max Count",
    "Retention Period",
    "Backup Cycle",
]

TABLE_SPEC_TITLE_KO = "테이블 명세서"
TABLE_SPEC_TITLE_EN = "Table Specification"
