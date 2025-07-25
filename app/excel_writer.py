import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.const import (
    BASE_COL,
    BASE_ROW,
    BOLD_FONT,
    BOLD_LARGE_FONT,
    BORDER_THIN,
    CENTER_ALIGN,
    COLUMN_HEADERS_EN,
    COLUMN_HEADERS_KO,
    COLUMN_WIDTHS,
    FILL_GRAY,
    FILL_HEADER,
    FILL_YELLOW,
    INDEX_HEADERS_EN,
    INDEX_HEADERS_KO,
    LEFT_ALIGN,
    META_FIELDS_EN,
    META_FIELDS_KO,
    TABLE_SPEC_TITLE_EN,
    TABLE_SPEC_TITLE_KO,
)
from app.utils import merge_and_style, set_row_style


def extract_length_from_type(type_str: str) -> str:
    """타입 문자열에서 길이 추출"""
    match = re.search(r"\((\d+)\)", type_str)
    return match.group(1) if match else "-"


def write_title(ws, row_idx, lang):
    """1. 시트 Title (테이블 명세서)"""
    title = TABLE_SPEC_TITLE_KO if lang == "ko" else TABLE_SPEC_TITLE_EN
    merge_and_style(
        ws,
        row_idx,
        BASE_COL,
        BASE_COL + 11,
        title,
        font=BOLD_LARGE_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    return row_idx + 1


def write_meta(ws, table_spec, row_idx, lang):
    """2. 상단 메타정보 영역 작성"""
    meta_fields = META_FIELDS_KO if lang == "ko" else META_FIELDS_EN
    meta_values = [""] * 7 + [table_spec["table_name"], table_spec["table_comment"]]
    for meta_idx, (meta_field, meta_value) in enumerate(
        zip(meta_fields, meta_values, strict=False)
    ):
        merge_and_style(
            ws,
            row_idx,
            BASE_COL,
            BASE_COL + 1,
            meta_field,
            font=BOLD_FONT,
            fill=FILL_YELLOW if 3 <= meta_idx <= 8 else FILL_GRAY,
            align=LEFT_ALIGN,
            border=BORDER_THIN,
        )
        merge_and_style(
            ws,
            row_idx,
            BASE_COL + 2,
            BASE_COL + 11,
            meta_value,
            align=LEFT_ALIGN,
            border=BORDER_THIN,
        )
        row_idx += 1
    return row_idx


def write_data_period(ws, row_idx):
    """3. 데이터 건수/주기 등 병합/헤더"""
    # 헤더
    merge_and_style(
        ws,
        row_idx,
        BASE_COL,
        BASE_COL + 2,
        "초기 예상 건수",
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 3,
        BASE_COL + 4,
        "증가 예상 건수",
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 5,
        BASE_COL + 6,
        "최대 예상 건수",
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 7,
        BASE_COL + 9,
        "data 보존 기간",
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 10,
        BASE_COL + 11,
        "data 백업 주기",
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
    )
    row_idx += 1
    # 값 라인
    merge_and_style(
        ws, row_idx, BASE_COL, BASE_COL + 2, "", align=LEFT_ALIGN, border=BORDER_THIN
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 3,
        BASE_COL + 4,
        "",
        align=LEFT_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 5,
        BASE_COL + 6,
        "",
        align=LEFT_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 7,
        BASE_COL + 9,
        "5년",
        align=LEFT_ALIGN,
        border=BORDER_THIN,
    )
    merge_and_style(
        ws,
        row_idx,
        BASE_COL + 10,
        BASE_COL + 11,
        "1년",
        align=LEFT_ALIGN,
        border=BORDER_THIN,
    )
    set_row_style(
        ws,
        row_idx,
        border=BORDER_THIN,
        align=LEFT_ALIGN,
        col_from=BASE_COL,
        col_to=BASE_COL + 11,
    )
    return row_idx + 1


def write_column_headers(ws, row_idx, lang):
    """4. 컬럼 스펙 테이블 헤더 (설명 병합)"""
    column_headers = COLUMN_HEADERS_KO if lang == "ko" else COLUMN_HEADERS_EN
    ws.append([""] * (BASE_COL - 1) + column_headers)
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 7,
        end_row=row_idx,
        end_column=BASE_COL + 9,
    )
    ws.cell(row_idx, BASE_COL + 7).value = "정의/설명"
    set_row_style(
        ws,
        row_idx,
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
        col_from=BASE_COL,
        col_to=BASE_COL + 11,
    )
    return row_idx + 1


def write_columns(ws, table_spec, row_idx):
    """5. 컬럼 데이터 (정의/설명 3칸 병합)"""
    for col_idx, column_spec in enumerate(table_spec["columns"], 1):
        ws.cell(row_idx, BASE_COL + 0, col_idx)
        ws.cell(row_idx, BASE_COL + 1, column_spec["column_name"])
        ws.cell(row_idx, BASE_COL + 2, column_spec["type"].upper())
        ws.cell(row_idx, BASE_COL + 3, extract_length_from_type(column_spec["type"]))
        ws.cell(row_idx, BASE_COL + 4, "Y" if column_spec.get("pk") else "N")
        ws.cell(row_idx, BASE_COL + 5, "Y" if column_spec.get("nn") else "N")
        ws.cell(row_idx, BASE_COL + 6, column_spec.get("default") or "-")
        ws.merge_cells(
            start_row=row_idx,
            start_column=BASE_COL + 7,
            end_row=row_idx,
            end_column=BASE_COL + 9,
        )
        ws.cell(row_idx, BASE_COL + 7, column_spec.get("comment", ""))
        ref_info = "-"
        if column_spec.get("ref_table") and column_spec.get("ref_column"):
            ref_info = f"{column_spec['ref_table']}.{column_spec['ref_column']}"
        ws.cell(row_idx, BASE_COL + 10, ref_info)
        ws.cell(row_idx, BASE_COL + 11, "-")
        for col_num in range(BASE_COL, BASE_COL + 12):
            ws.cell(row_idx, col_num).alignment = (
                CENTER_ALIGN
                if col_num in [BASE_COL, BASE_COL + 4, BASE_COL + 5]
                else LEFT_ALIGN
            )
            ws.cell(row_idx, col_num).border = BORDER_THIN
        row_idx += 1
    return row_idx


def write_index(ws, row_idx, lang):
    """
    6. Index 스펙 (헤더/PK 인덱스 예시)
    - 인덱스 헤더 라인, PK 인덱스(예시) 데이터, 병합 및 스타일 적용
    """
    index_headers = INDEX_HEADERS_KO if lang == "ko" else INDEX_HEADERS_EN
    ws.append([""] * (BASE_COL - 1) + index_headers)
    # 헤더 병합
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 1,
        end_row=row_idx,
        end_column=BASE_COL + 3,
    )  # Index name
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 4,
        end_row=row_idx,
        end_column=BASE_COL + 6,
    )  # Index type
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 7,
        end_row=row_idx,
        end_column=BASE_COL + 9,
    )  # Unique
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 10,
        end_row=row_idx,
        end_column=BASE_COL + 11,
    )  # 구성 컬럼
    set_row_style(
        ws,
        row_idx,
        font=BOLD_FONT,
        fill=FILL_HEADER,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
        col_from=BASE_COL,
        col_to=BASE_COL + 11,
    )
    row_idx += 1

    # PK 인덱스 데이터 예시
    index_data = [1, "-", "", "", "-", "", "", "-", "", "", "", "-", ""]
    ws.append([""] * (BASE_COL - 1) + index_data)
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 1,
        end_row=row_idx,
        end_column=BASE_COL + 3,
    )
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 4,
        end_row=row_idx,
        end_column=BASE_COL + 6,
    )
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 7,
        end_row=row_idx,
        end_column=BASE_COL + 9,
    )
    ws.merge_cells(
        start_row=row_idx,
        start_column=BASE_COL + 10,
        end_row=row_idx,
        end_column=BASE_COL + 11,
    )
    set_row_style(
        ws,
        row_idx,
        align=CENTER_ALIGN,
        border=BORDER_THIN,
        col_from=BASE_COL,
        col_to=BASE_COL + 11,
    )
    row_idx += 1
    return row_idx


def set_column_widths(ws):
    """엑셀 시트 열 너비 조정"""
    for col_num, width in enumerate(COLUMN_WIDTHS, BASE_COL):
        ws.column_dimensions[get_column_letter(col_num)].width = width


def write_table_sheet(ws, table_spec, lang):
    """
    테이블 단위 시트 작성 (모든 블록 호출)
    """
    row = BASE_ROW
    row = write_title(ws, row, lang)
    row = write_meta(ws, table_spec, row, lang)
    row = write_data_period(ws, row)
    row = write_column_headers(ws, row, lang)
    row = write_columns(ws, table_spec, row)
    row = write_index(ws, row, lang)
    set_column_widths(ws)


def write_excel_spec(
    table_spec_dict: dict[str, Any], output_excel_path: Path, lang: str
):
    """
    메인: 전체 엑셀 파일 생성, 각 시트 작성
    """
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, table_list in table_spec_dict.items():
        for table_spec in table_list:
            ws = wb.create_sheet(title=f"{sheet_name}_{table_spec['table_name']}"[:31])
            ws.sheet_view.zoomScale = 85
            write_table_sheet(ws, table_spec, lang=lang)
    wb.save(output_excel_path)
